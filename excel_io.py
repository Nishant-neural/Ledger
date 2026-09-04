"""
excel_io.py - Import transactions from an Excel sheet and export the ledger to Excel.
Uses openpyxl only (no pandas dependency needed).
"""
import openpyxl
from datetime import datetime
import database as db

# Column names we try to auto-detect in the uploaded Excel sheet.
# The matching is case-insensitive and flexible on naming.
COLUMN_ALIASES = {
    "date": ["date", "transaction date", "txn date"],
    "party": ["party", "name", "customer", "vendor", "payee", "client"],
    "category": ["category", "type", "head", "account"],
    "description": ["description", "details", "narration", "remarks", "note"],
    "debit": ["debit", "expense", "paid", "out", "withdrawal"],
    "credit": ["credit", "income", "received", "in", "deposit"],
    "item": ["item", "product", "item name", "product name"],
    "stock": ["stock", "stock/qty", "quantity", "qty", "units", "sold"],
    "buying_rate": ["buying rate", "buy rate", "purchase rate", "cost price"],
    "selling_rate": ["selling rate", "sell rate", "sale rate", "selling price"],
}


def _normalize(text):
    return str(text).strip().lower() if text is not None else ""


def detect_columns(header_row):
    """
    Given the header row (list of cell values), figure out which column index
    maps to which field (date, party, category, description, debit, credit).
    Returns a dict: field_name -> column_index (0-based), or None if not found.
    """
    mapping = {}
    normalized_headers = [_normalize(h) for h in header_row]
    for field, aliases in COLUMN_ALIASES.items():
        found_idx = None
        for idx, header in enumerate(normalized_headers):
            if header in aliases:
                found_idx = idx
                break
        mapping[field] = found_idx
    return mapping


def read_excel_preview(filepath, max_rows=5):
    """Read the header row and a few preview rows for user confirmation."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = list(rows[0])
    preview = [list(r) for r in rows[1:1 + max_rows]]
    return header, preview


def import_excel(filepath, column_mapping=None):
    """
    Import transactions from an Excel file into the ledger database.
    column_mapping: optional dict overriding auto-detected columns,
                    e.g. {"date": 0, "party": 1, "debit": 4, "credit": 5}
    Returns (success_count, error_count, error_messages)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0, ["The sheet appears to be empty."]

    header = list(rows[0])
    data_rows = rows[1:]

    mapping = column_mapping if column_mapping else detect_columns(header)

    if mapping.get("date") is None:
        return 0, 0, ["Could not find a 'Date' column. Please map columns manually."]

    success, errors = 0, []
    for i, row in enumerate(data_rows, start=2):  # row 2 = first data row (1-indexed in Excel)
        try:
            date_val = row[mapping["date"]] if mapping.get("date") is not None else None
            if date_val is None or str(date_val).strip() == "":
                continue  # skip blank rows

            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val).strip()

            party = str(row[mapping["party"]]).strip() if mapping.get("party") is not None and row[mapping["party"]] is not None else ""
            category = str(row[mapping["category"]]).strip() if mapping.get("category") is not None and row[mapping["category"]] is not None else "Misc Expense"
            description = str(row[mapping["description"]]).strip() if mapping.get("description") is not None and row[mapping["description"]] is not None else ""

            debit = row[mapping["debit"]] if mapping.get("debit") is not None else 0
            credit = row[mapping["credit"]] if mapping.get("credit") is not None else 0
            debit = float(debit) if debit not in (None, "") else 0.0
            credit = float(credit) if credit not in (None, "") else 0.0
            item = str(row[mapping["item"]]).strip() if mapping.get("item") is not None and row[mapping["item"]] is not None else ""
            stock = row[mapping["stock"]] if mapping.get("stock") is not None else 0
            buying_rate = row[mapping["buying_rate"]] if mapping.get("buying_rate") is not None else 0
            selling_rate = row[mapping["selling_rate"]] if mapping.get("selling_rate") is not None else 0
            stock = float(stock) if stock not in (None, "") else 0.0
            buying_rate = float(buying_rate) if buying_rate not in (None, "") else 0.0
            selling_rate = float(selling_rate) if selling_rate not in (None, "") else 0.0

            db.add_transaction(date_str, party, category, description, debit, credit,
                               item, stock, buying_rate, selling_rate)
            if category:
                db.add_category(category)
            success += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    return success, len(errors), errors


def export_excel(filepath, rows):
    """Export transactions and SME helper reports to an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    headers = ["Date", "Item", "Party", "Category", "Stock/Qty", "Buying Rate",
               "Selling Rate", "Total Sales", "Description", "Debit", "Credit", "Balance"]
    ws.append(headers)

    running_balance = 0.0
    for r in rows:
        running_balance += (r["credit"] or 0) - (r["debit"] or 0)
        ws.append([
            r["date"], r["item"], r["party"], r["category"], r["stock"],
            r["buying_rate"], r["selling_rate"],
            round((r["stock"] or 0) * (r["selling_rate"] or 0), 2),
            r["description"], r["debit"], r["credit"], round(running_balance, 2)
        ])

    # basic column widths
    widths = [12, 18, 20, 16, 10, 12, 12, 12, 30, 10, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    cashbook_ws = wb.create_sheet("Daily Cashbook")
    cashbook_ws.append(["Date", "Cash In", "Cash Out", "Net Change", "Running Balance"])
    for row in db.get_cashbook():
        cashbook_ws.append([
            row["date"], row["cash_in"], row["cash_out"],
            row["net"], row["running_balance"]
        ])

    party_ws = wb.create_sheet("Party Balances")
    party_ws.append(["Party", "Total Credit", "Total Debit", "Net Balance", "Status"])
    for row in db.get_party_balances():
        balance = row["net_balance"] or 0
        status = "Receivable" if balance > 0 else "Payable" if balance < 0 else "Settled"
        party_ws.append([
            row["party"], row["total_credit"], row["total_debit"], balance, status
        ])

    inventory_ws = wb.create_sheet("Inventory")
    inventory_ws.append([
        "Item", "Purchased Qty", "Sold Qty", "Available Qty", "Avg Cost",
        "Stock Value", "Sales Value", "Gross Profit"
    ])
    for row in db.get_inventory_report():
        inventory_ws.append([
            row["item"], row["purchased_qty"], row["sold_qty"], row["available_qty"],
            row["avg_cost"], row["stock_value"], row["sales_value"], row["gross_profit"]
        ])

    low_stock_ws = wb.create_sheet("Low Stock")
    low_stock_ws.append(["Item", "Available Qty", "Sold Qty", "Stock Value", "Status"])
    for row in db.get_low_stock_items(limit=5):
        status = "Out of stock" if row["available_qty"] <= 0 else "Low stock"
        low_stock_ws.append([
            row["item"], row["available_qty"], row["sold_qty"], row["stock_value"], status
        ])

    for sheet in wb.worksheets[1:]:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    wb.save(filepath)
